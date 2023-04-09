# solved python search path for import module in other sub directories
import sys
import os
current_working_directory = os.getcwd()
sys.path.append(current_working_directory)

import pandas as pd
import script_modules.subscript.line.CreateLine as line
from pandas import ExcelWriter
from script_modules.subscript.cleanup_file.cleanup import DataCleaner
import openpyxl

# show full pandas df
pd.set_option('display.max_rows', None)

############################################################################

class WriteData:

    def __init__(self, dateObj ,path ,title = '', formula = {}, formula_header = {}) :
        self.path = path
        self.title = title
        self.formula = formula
        self.formula_header = formula_header
        self.dateObj = dateObj
        self.type = type

    def write_date_to_excel(self, list_dfs, list_of_sheet_names):

        line.title_between_line(f"{self.title}")

        def remove_empty_rows(sheet, row):
            # iterate the row object
            for cell in row:
                # check the value of each cell in
                # the row, if any of the value is not
                # None return without removing the row
                if cell.value != None:
                    return
            # get the row number from the first cell
            # and remove the row
            sheet.delete_rows(row[0].row, 1)

        with ExcelWriter(self.path) as writer:
            n = 0
            for df in list_dfs:
                sheetName = list_of_sheet_names[n]

                # check formula
                if len(self.formula) > 0:
                    
                    formula_items = iter(self.formula[n].items())
                    while key_value := next(formula_items, None):
                        formula_temp = {key_value[0] : key_value[1]}
                        formula_header_tmp = self.formula_header[n]


                        for a, b in zip(formula_temp, formula_header_tmp):

                            formula_column = a
                            formula = self.formula[n][a]

                            # formula_header_column = b
                            formula_header = self.formula_header[n][b]
                            
                            j = 0
                            k = 3
                            while j < len(df):
                                try:
                                    df.loc[j:j, pd.IndexSlice[:, formula_column]] = formula.replace('$', f'{k}')
                                except Exception as e:
                                    print(e)
                                j += 1
                                k += 1

                        # find type for update header
                        if 'file' in sheetName.lower():
                            type = 'file'
                        elif 'db' in sheetName.lower():
                            type = 'db'

                        headerObj = DataCleaner(self.dateObj, df, type)
                        df = headerObj.update_header(formula_header)
                
                elif len(self.formula) <= 0:
                    pass
                  
                # write to excel
                df.to_excel(writer, list_of_sheet_names[n])
                n += 1


        # remove empty rows
        for anySheet in list_of_sheet_names:
            wb = openpyxl.load_workbook(self.path)
            sheet = wb[anySheet]
            # print("Maximum rows before removing:", sheet.max_row)
            # iterate the sheet object
            for row in sheet:
                remove_empty_rows(sheet,row)
            # print("Maximum rows after removing:",sheet.max_row)
            # save the file to the path
            wb.save(self.path)


        # go to next df




        file_name_tmp = self.path.split("\\")
        file_name = file_name_tmp[-1]

        print(f"  {file_name} saved successfully .\n")


    def write_date_to_csv(self, df):

        df.to_csv(self.path, index=False)

        file_name_tmp = self.path.split("\\")
        file_name = file_name_tmp[-1]

        print(f"  {file_name} saved successfully .\n")